"""Export scraped data to CSV and Excel."""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")


def to_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def to_excel(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Books", startrow=1)
        wb = writer.book
        ws = wb["Books"]

        # Title row
        ws["A1"] = f"Books Scraped: {len(df)} titles"
        ws["A1"].font = Font(size=13, bold=True, color="1B4F72")

        # Style header row (row 2)
        for cell in ws[2]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Alternate row shading
        for row_idx in range(3, len(df) + 3):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = ALT_FILL

        # Auto column widths
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            length = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length, 50)

        # Freeze header and add filter
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}2"

        # Summary sheet
        ws_sum = wb.create_sheet("Summary")
        ws_sum["A1"] = "Scrape Summary"
        ws_sum["A1"].font = Font(size=13, bold=True, color="1B4F72")

        summary_data = {
            "Total Books": len(df),
            "Avg Price (£)": round(df["price_gbp"].mean(), 2),
            "Min Price (£)": df["price_gbp"].min(),
            "Max Price (£)": df["price_gbp"].max(),
            "Avg Rating": round(df["rating"].mean(), 2),
            "In Stock": (df["availability"] == "In stock").sum(),
        }
        for i, (k, v) in enumerate(summary_data.items(), start=3):
            ws_sum[f"A{i}"] = k
            ws_sum[f"B{i}"] = v
            ws_sum[f"A{i}"].font = Font(bold=True)

        # Rating breakdown
        ws_sum["D2"] = "Rating Breakdown"
        ws_sum["D2"].font = Font(bold=True)
        ws_sum["D3"] = "Stars"
        ws_sum["E3"] = "Count"
        for j, (stars, count) in enumerate(
            df["rating"].value_counts().sort_index().items(), start=4
        ):
            ws_sum[f"D{j}"] = f"{'★' * stars}"
            ws_sum[f"E{j}"] = count
